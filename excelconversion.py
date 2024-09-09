import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import PatternFill

xml_file_path = "employee_data.xml"

# Parse the XML file
try:
    tree = ET.parse(xml_file_path)
    root = tree.getroot()
except ET.ParseError as e:
    print(f"Error parsing XML file: {e}")
    exit(1)

wb = Workbook()
ws = wb.active
ws.title = "Employees"

header = ['id', 'name', 'position', 'department']
ws.append(header)

# Iterate over each employee element in the XML and write data to Excel
for employee in root.findall('.//employee'):
    emp_id = employee.find('id').text
    name = employee.find('name').text
    position = employee.find('position').text
    department = employee.find('department').text
    ws.append([emp_id, name, position, department])

header_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
for cell in ws[1]:
    cell.fill = header_fill

for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter  # Get the column name
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column_letter].width = adjusted_width

wb.save('employees.xlsx')
